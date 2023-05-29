import os
from flask import Flask,render_template,request
import win32com.client as win32
import datetime as dt
from pathlib import Path
import glob
import pandas as pd
import xlrd
import openpyxl
from datetime import timedelta
import math

app=Flask(__name__)

@app.route('/',methods=['GET','POST'])
def countfiles():
    if request.method == 'POST':
        date=request.form.get('date')
        win32.pythoncom.CoInitialize()
        date_str = date+" "+"00:00"
        date = dt.datetime.strptime(date_str,"%m-%d-%Y %H:%M")
        year,month,day=map(str,[date.year,date.month,date.day])
        outlook = win32.Dispatch("Outlook.Application").GetNameSpace("MAPI")

        # selects the inbox folder
        # for more info -> https://learn.microsoft.com/en-us/office/vba/api/outlook.oldefaultfolders
        inbox = outlook.GetDefaultfolder(6)

        # subtract one day from date, because the forecast for any given day comes in from hour 22 of the day prior to hour 22 of the day of
        date_sub1 = date - dt.timedelta(days=1)

        # extract the day
        sub1_day = date_sub1.day

        # replaces the start and end time
        dday = int(day)
        start_time = date_sub1.replace(day=sub1_day, hour=22, minute=0, second=0).strftime('%m/%d/%Y %H:%M %p')
        end_time = date.replace(day=dday, hour=21, minute=59, second=59).strftime('%m/%d/%Y %H:%M %p')

        # filters emails in inbox folder using received time and sender email address
        filter_str = "[SenderEmailAddress] = 'reportingservices@NorpacPaper.com'"
        filter_str += " AND [ReceivedTime] >= '" + start_time + "' And [ReceivedTime] <= '" + end_time + "'"
        messages =inbox.Items.Restrict(filter_str)

        # check if any messages were found
        if messages.Count == 0:
            print(f"No forecasts are found for {date.date()}!")
            return
        else:
            # creates a path for the attachments to be saved in
            path = Path(r'C:\Norpac') / year / month / day
            if path.exists():
                pass
            else:
                # creates directory of the missing parents of the path
                path.mkdir(parents=True)

            # go through all the filtered messages and download the forecasts
            try:
                count = 0
                for message in list(messages):
                    try:
                        # downloads attachments to the directory above
                        for attachment in message.Attachments:
                            attachment.SaveASFile(str(path / attachment.FileName))
                            count += 1  # increment counter variable on successful download
                    # throws an exception object if error happens downloading the file
                    except Exception as e:
                        print("error saving the attachment:" + str(e))

                
            # throws an execption object if error happens while processing filtered emails
            except Exception as e:
                print("error when processing email messages:" + str(e))

            excel_files=glob.glob(os.path.join(path,'*xls'))

            while True:
                if os.path.exists(r"C:\Automate\cowlitz_directory.txt"):
                    with open(r"C:\Automate\cowlitz_directory.txt", "r") as f:
                        cowlitz_directory = f.read().strip()
                        break
            else:
                print(message="\nCould not locate a path for Cowlitz Large industry report.xlsx file!")

            # Define the file path and sheet name of the existing Excel file
        cowl_sheet_path = fr"{cowlitz_directory}"
        sheet_name = "Norpac Load"

        # Load the Cowlitz industry report workbook file using pandas
        cs_df = pd.read_excel(cowl_sheet_path, sheet_name=sheet_name)

        # dates on column 1 of Cowlitz
        forecast_dates = cs_df.iloc[:, 2]

        forecast_dates_list = forecast_dates.tolist()

        first_row = cs_df.iloc[[0], 2:]

        # load the hours row
        forecast_hours = pd.DataFrame(first_row.columns)
        forecast_hours = forecast_hours[3:]

        # Load the Cowlitz industry report workbook using openpyxl
        book = openpyxl.load_workbook(cowl_sheet_path)
        # Get the target sheet by name
        sheet = book[sheet_name]

        j = 0
        for file in excel_files:
            # Read the forecast Excel file into a DataFrame
            wb = xlrd.open_workbook(file, logfile=open(os.devnull, 'w'))  # supress warning
            fc_df = pd.read_excel(wb)

            # Extract datetime of when forecast was conducted
            fc_datetime_str = str(fc_df.iloc[1, 0])
            # reformat datetime from (yyyy-mm-dd HH:MM:SS.ff to mm/dd/yyyy HH:MM)
            fc_datetime_obj = date_convert(fc_datetime_str)
            # remove 0s from start of day and month and hour
            fc_datetime = '{}/{}/{} {}:{:02d}'.format(fc_datetime_obj.month, fc_datetime_obj.day, fc_datetime_obj.year, fc_datetime_obj.hour, fc_datetime_obj.minute)

            # save the forecast values from forecast row as df
            forecast_raw = fc_df.iloc[3, 2:27]
            # convert df to a list and round the values to 2 decimal points
            forecast = [round(i, 2) for i in forecast_raw]
            # remove NaN values from forecast list
            filtered_forecast = list(filter(lambda x: not math.isnan(x), forecast))

            # find the first none NaN value and its index in the forecast list and set it as realtime forecast
            for i, x in enumerate(forecast):
                if not math.isnan(x):
                    rt_idx = i
                    real_time_forecast = x
                    break
                else:
                    rt_idx = 0
                    real_time_forecast = 0


            # Extract the date of the target day for the forecast
            # and add its srating index to its datetime, because if forecast starts at index 6 in the list
            # it means that this forecast starts from hour ending 7
            target_date_str = str(fc_df.iloc[3, 0])
            # reformat datetime from yyyy-mm-dd 00:00:00 to mm/dd/yyyy 01:00:00
            target_datetime_obj = date_convert(target_date_str)
            target_datetime_obj2 = target_datetime_obj + timedelta(hours=rt_idx + 1) # Add index + 1 to dt
            target_datetime = target_datetime_obj2.strftime("%m/%d/%Y %H:%M")

            # column to start writing from for every hour of forecast, first not-empty column then next
            start_col = forecast_hours.index[-1] + 4 + j
            # row to start writing from
            start_row = forecast_dates[forecast_dates == target_datetime].index[0] + 2 # finds the row with matching date in datetime column and the forecast

            # paste the realtime forecast value into Real-Time column
            sheet.cell(row=start_row, column=2).value = real_time_forecast

            # paste the forecast datetime the header row
            sheet.cell(row=1, column=start_col).value = fc_datetime

            # paste the hourly forecasts values vertically in every column
            for index, value in enumerate(filtered_forecast):
                sheet.cell(row=start_row + index, column=start_col).value = value

            j += 1

        # write the entire Datetime column everytime
        # this needs to be done because openpyxl fails to save the cells that populate as a result of excel formula formatting
        # more info: https://stackoverflow.com/questions/74766097/python-excel-formula-file-saving-issues-with-openpyxl
        for i in range(len(forecast_dates_list)):
            sheet.cell(row=2 + i, column=3).value = forecast_dates_list[i]

        # Save the changes to the workbook
        book.save(cowl_sheet_path)
    
        return render_template('result.html')
    return render_template('index.html')


def date_convert(fc_datetime):
    # check to see if the datetime has fraction of second
    if '.' in fc_datetime:
        # Timestamp has milliseconds
        dt_obj = dt.datetime.strptime(fc_datetime, '%Y-%m-%d %H:%M:%S.%f')
    else:
        # Timestamp has no milliseconds
        dt_obj = dt.datetime.strptime(fc_datetime, '%Y-%m-%d %H:%M:%S')
    return dt_obj


if __name__ == '__main__':
    app.run()
                           